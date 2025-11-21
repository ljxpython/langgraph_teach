"""
UI自动化测试工具集
包含Excel保存、HTML报告生成、测试用例管理等功能
"""
import asyncio
import base64
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from langchain.agents import create_agent
from langchain_core.tools import tool
from langchain_mcp_adapters.client import MultiServerMCPClient
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from src.llms import get_default_model
model = get_default_model()



# 全局输出目录
OUTPUT_DIR = os.path.join(os.getcwd(), "ui_test_reports")


def ensure_output_dir():
    """确保输出目录存在"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    return OUTPUT_DIR


@tool
def save_test_cases_to_excel(test_cases: str, filename: str = None) -> str:
    """
    将测试用例保存到Excel文件

    Args:
        test_cases: JSON格式的测试用例列表，每个用例包含：
                   - case_id: 用例ID
                   - title: 用例标题
                   - steps: 测试步骤
                   - expected: 预期结果
                   - actual: 实际结果（可选）
                   - status: 状态（Pass/Fail/Skip）
                   - screenshot: 截图base64（可选）
        filename: 文件名（可选，默认使用时间戳）

    Returns:
        保存的文件路径
    """
    try:
        # 解析测试用例
        if isinstance(test_cases, str):
            cases = json.loads(test_cases)
        else:
            cases = test_cases

        if not isinstance(cases, list):
            cases = [cases]

        # 创建输出目录
        output_dir = ensure_output_dir()

        # 生成文件名
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"test_cases_{timestamp}.xlsx"

        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        filepath = os.path.join(output_dir, filename)

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"

        # 设置表头样式
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 写入表头
        headers = [
            "用例ID",
            "用例标题",
            "测试步骤",
            "预期结果",
            "实际结果",
            "状态",
            "备注",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # 写入数据
        for row_idx, case in enumerate(cases, 2):
            # 用例ID
            ws.cell(
                row=row_idx, column=1, value=case.get("case_id", f"TC_{row_idx-1:03d}")
            )
            # 用例标题
            ws.cell(row=row_idx, column=2, value=case.get("title", ""))
            # 测试步骤
            steps = case.get("steps", "")
            if isinstance(steps, list):
                steps = "\n".join(f"{i+1}. {step}" for i, step in enumerate(steps))
            ws.cell(row=row_idx, column=3, value=steps)
            # 预期结果
            ws.cell(row=row_idx, column=4, value=case.get("expected", ""))
            # 实际结果
            ws.cell(row=row_idx, column=5, value=case.get("actual", ""))
            # 状态
            status = case.get("status", "Not Run")
            status_cell = ws.cell(row=row_idx, column=6, value=status)

            # 根据状态设置颜色
            if status == "Pass":
                status_cell.fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
                status_cell.font = Font(color="006100")
            elif status == "Fail":
                status_cell.fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )
                status_cell.font = Font(color="9C0006")
            elif status == "Skip":
                status_cell.fill = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                )
                status_cell.font = Font(color="9C6500")

            # 备注
            ws.cell(row=row_idx, column=7, value=case.get("remarks", ""))

            # 应用边框和对齐
            for col in range(1, 8):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # 调整列宽
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 20

        # 保存文件
        wb.save(filepath)

        return f"测试用例已保存到: {filepath}"

    except Exception as e:
        return f"保存Excel失败: {str(e)}"


@tool
def generate_html_report(
    test_results: str, report_title: str = "UI自动化测试报告"
) -> str:
    """
    生成图文并茂的HTML测试报告

    Args:
        test_results: JSON格式的测试结果，包含：
                     - summary: 测试摘要（total, pass, fail, skip）
                     - test_cases: 测试用例列表
                     - screenshots: 截图列表（可选）
                     - start_time: 开始时间
                     - end_time: 结束时间
        report_title: 报告标题

    Returns:
        生成的HTML报告路径
    """
    try:
        # 解析测试结果
        if isinstance(test_results, str):
            results = json.loads(test_results)
        else:
            results = test_results

        # 创建输出目录
        output_dir = ensure_output_dir()

        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"test_report_{timestamp}.html"
        filepath = os.path.join(output_dir, filename)

        # 获取测试摘要
        summary = results.get("summary", {})
        total = summary.get("total", 0)
        passed = summary.get("pass", 0)
        failed = summary.get("fail", 0)
        skipped = summary.get("skip", 0)
        pass_rate = (passed / total * 100) if total > 0 else 0

        # 获取测试用例
        test_cases = results.get("test_cases", [])

        # 获取时间信息
        start_time = results.get(
            "start_time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        end_time = results.get("end_time", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        # 生成HTML内容
        html_content = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{report_title}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 10px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        .header h1 {{
            font-size: 32px;
            margin-bottom: 10px;
        }}
        .header p {{
            font-size: 14px;
            opacity: 0.9;
        }}
        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            padding: 30px;
            background: #f8f9fa;
        }}
        .summary-card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            text-align: center;
        }}
        .summary-card h3 {{
            font-size: 14px;
            color: #666;
            margin-bottom: 10px;
            text-transform: uppercase;
        }}
        .summary-card .value {{
            font-size: 36px;
            font-weight: bold;
            margin-bottom: 5px;
        }}
        .summary-card.total .value {{ color: #667eea; }}
        .summary-card.pass .value {{ color: #28a745; }}
        .summary-card.fail .value {{ color: #dc3545; }}
        .summary-card.skip .value {{ color: #ffc107; }}
        .summary-card.rate .value {{ color: #17a2b8; }}
        .test-cases {{
            padding: 30px;
        }}
        .test-case {{
            background: white;
            border: 1px solid #e0e0e0;
            border-radius: 8px;
            margin-bottom: 20px;
            overflow: hidden;
        }}
        .test-case-header {{
            padding: 15px 20px;
            background: #f8f9fa;
            border-bottom: 1px solid #e0e0e0;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        .test-case-title {{
            font-size: 16px;
            font-weight: bold;
            color: #333;
        }}
        .test-case-status {{
            padding: 5px 15px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: bold;
            text-transform: uppercase;
        }}
        .status-pass {{
            background: #d4edda;
            color: #155724;
        }}
        .status-fail {{
            background: #f8d7da;
            color: #721c24;
        }}
        .status-skip {{
            background: #fff3cd;
            color: #856404;
        }}
        .test-case-body {{
            padding: 20px;
        }}
        .test-case-section {{
            margin-bottom: 15px;
        }}
        .test-case-section h4 {{
            font-size: 14px;
            color: #666;
            margin-bottom: 8px;
            text-transform: uppercase;
        }}
        .test-case-section p, .test-case-section ol {{
            font-size: 14px;
            color: #333;
            line-height: 1.6;
        }}
        .test-case-section ol {{
            padding-left: 20px;
        }}
        .screenshot {{
            margin-top: 15px;
            text-align: center;
        }}
        .screenshot img {{
            max-width: 100%;
            border-radius: 8px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        }}
        .footer {{
            background: #f8f9fa;
            padding: 20px;
            text-align: center;
            color: #666;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{report_title}</h1>
            <p>生成时间: {end_time}</p>
        </div>

        <div class="summary">
            <div class="summary-card total">
                <h3>总用例数</h3>
                <div class="value">{total}</div>
            </div>
            <div class="summary-card pass">
                <h3>通过</h3>
                <div class="value">{passed}</div>
            </div>
            <div class="summary-card fail">
                <h3>失败</h3>
                <div class="value">{failed}</div>
            </div>
            <div class="summary-card skip">
                <h3>跳过</h3>
                <div class="value">{skipped}</div>
            </div>
            <div class="summary-card rate">
                <h3>通过率</h3>
                <div class="value">{pass_rate:.1f}%</div>
            </div>
        </div>

        <div class="test-cases">
            <h2 style="margin-bottom: 20px; color: #333;">测试用例详情</h2>
"""

        # 添加测试用例
        for case in test_cases:
            case_id = case.get("case_id", "N/A")
            title = case.get("title", "未命名用例")
            status = case.get("status", "Not Run")
            steps = case.get("steps", [])
            expected = case.get("expected", "")
            actual = case.get("actual", "")
            screenshot = case.get("screenshot", "")

            status_class = f"status-{status.lower()}"

            # 处理步骤
            steps_html = ""
            if isinstance(steps, list):
                steps_html = (
                    "<ol>" + "".join(f"<li>{step}</li>" for step in steps) + "</ol>"
                )
            else:
                steps_html = f"<p>{steps}</p>"

            html_content += f"""
            <div class="test-case">
                <div class="test-case-header">
                    <div class="test-case-title">{case_id}: {title}</div>
                    <div class="test-case-status {status_class}">{status}</div>
                </div>
                <div class="test-case-body">
                    <div class="test-case-section">
                        <h4>测试步骤</h4>
                        {steps_html}
                    </div>
                    <div class="test-case-section">
                        <h4>预期结果</h4>
                        <p>{expected}</p>
                    </div>
                    <div class="test-case-section">
                        <h4>实际结果</h4>
                        <p>{actual}</p>
                    </div>
"""

            # 添加截图
            if screenshot:
                # 如果是base64编码的图片
                if screenshot.startswith("data:image"):
                    img_src = screenshot
                elif screenshot.startswith("/") or screenshot.startswith("http"):
                    img_src = screenshot
                else:
                    img_src = f"data:image/png;base64,{screenshot}"

                html_content += f"""
                    <div class="screenshot">
                        <h4>截图</h4>
                        <img src="{img_src}" alt="测试截图">
                    </div>
"""

            html_content += """
                </div>
            </div>
"""

        html_content += f"""
        </div>

        <div class="footer">
            <p>测试开始时间: {start_time} | 测试结束时间: {end_time}</p>
            <p>由 LangGraph UI自动化测试系统 生成</p>
        </div>
    </div>
</body>
</html>
"""

        # 保存HTML文件
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html_content)

        return f"HTML测试报告已生成: {filepath}"

    except Exception as e:
        return f"生成HTML报告失败: {str(e)}"


@tool
def set_output_directory(directory: str) -> str:
    """
    设置测试报告输出目录

    Args:
        directory: 输出目录路径

    Returns:
        设置结果消息
    """
    global OUTPUT_DIR
    try:
        # 创建目录
        os.makedirs(directory, exist_ok=True)
        OUTPUT_DIR = directory
        return f"输出目录已设置为: {OUTPUT_DIR}"
    except Exception as e:
        return f"设置输出目录失败: {str(e)}"


@tool
def get_output_directory() -> str:
    """
    获取当前测试报告输出目录

    Returns:
        当前输出目录路径
    """
    return OUTPUT_DIR


def get_chrome_mcp_tools():
    client = MultiServerMCPClient(
        {
            "chrome_mcp": {
                "url": "http://127.0.0.1:12306/mcp",
                "transport": "streamable_http",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools


# UI自动化测试Agent
ui_test_agent = create_agent(
    model=model,
    tools=get_chrome_mcp_tools()
    + [
        save_test_cases_to_excel,
        generate_html_report,
        set_output_directory,
        get_output_directory,
    ],
    system_prompt="""你是一个专业的UI自动化测试专家。你的职责是：

1. **测试用例设计**：根据用户需求，设计清晰、完整的测试用例
   - 每个测试用例应包含：用例ID、标题、测试步骤、预期结果
   - 测试步骤要详细、可执行
   - 覆盖正常流程和异常场景

2. **自动化测试执行**：使用Chrome MCP工具执行测试
   - 使用chrome_navigate_chrome_mcp打开网页
   - 使用chrome_get_web_content_chrome_mcp获取页面内容
   - 使用chrome_click_element_chrome_mcp点击元素
   - 使用chrome_fill_or_select_chrome_mcp填写表单
   - 使用chrome_screenshot_chrome_mcp截取关键步骤的截图
   - 使用chrome_get_interactive_elements_chrome_mcp获取可交互元素

3. **测试结果记录**：
   - 记录每个测试用例的执行结果（Pass/Fail）
   - 记录实际结果与预期结果的对比
   - 保存关键步骤的截图（base64格式）

4. **报告生成**：
   - 使用save_test_cases_to_excel将测试用例保存到Excel
   - 使用generate_html_report生成图文并茂的HTML测试报告
   - 报告应包含：测试摘要、详细用例、截图、统计数据

5. **工作流程**：
   a. 理解用户需求，设计测试用例
   b. 打开目标网站
   c. 逐个执行测试用例，记录结果和截图
   d. 汇总测试结果
   e. 生成Excel和HTML报告
   f. 向用户报告测试完成情况和报告路径

注意事项：
- 每个测试步骤都要截图记录
- 遇到错误要详细记录错误信息
- 测试报告要清晰、专业、易读
- 所有文件保存到ui_test_reports目录
""",
)
